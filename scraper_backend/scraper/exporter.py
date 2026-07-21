import json
import logging
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class DataExporter:
    """
    Exports enriched perfume catalog data into Formatted Excel (.xlsx), CSV, and JSON files.
    """

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_json(self, data: List[Dict[str, Any]], filename: str = "lattafa_perfumes.json") -> Path:
        """Exports data to JSON format."""
        file_path = self.output_dir / filename
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logging.info(f"Exported JSON file to {file_path}")
        return file_path

    def export_csv(self, data: List[Dict[str, Any]], filename: str = "lattafa_perfumes.csv") -> Path:
        """Exports main product catalog to CSV format."""
        file_path = self.output_dir / filename
        
        # Flatten list of images and omit complex variants dict for clean CSV
        flat_data = []
        for item in data:
            row = item.copy()
            row.pop("variants", None)
            row.pop("all_images", None)
            flat_data.append(row)

        df = pd.DataFrame(flat_data)
        df.to_csv(file_path, index=False, encoding="utf-8-sig")
        logging.info(f"Exported CSV file to {file_path}")
        return file_path

    def export_excel(self, data: List[Dict[str, Any]], filename: str = "lattafa_perfumes_catalog.xlsx") -> Path:
        """
        Exports a multi-sheet, beautifully formatted Excel workbook.
        - Sheet 1: Perfumes Catalog
        - Sheet 2: Product Variants
        - Sheet 3: Catalog Summary & Analytics
        """
        file_path = self.output_dir / filename
        wb = openpyxl.Workbook()
        
        # Setup styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # ----------------------------------------------------
        # SHEET 1: PERFUMES CATALOG
        # ----------------------------------------------------
        ws_catalog = wb.active
        ws_catalog.title = "Perfumes Catalog"

        catalog_headers = [
            "ID", "Title", "Brand / Vendor", "Gender", "Concentration", "Volume",
            "Price ($)", "Compare At ($)", "Discount (%)", "Top Notes", "Heart Notes", "Base Notes",
            "In Stock", "Total Variants", "Main Image", "URL", "Created Date"
        ]
        
        ws_catalog.append(catalog_headers)

        for cell in ws_catalog[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for p in data:
            ws_catalog.append([
                p.get("id"),
                p.get("title"),
                p.get("vendor"),
                p.get("gender"),
                p.get("concentration"),
                p.get("volume"),
                p.get("min_price"),
                p.get("compare_at_price") or "",
                p.get("discount_percent"),
                p.get("top_notes"),
                p.get("heart_notes"),
                p.get("base_notes"),
                "YES" if p.get("is_available") else "NO",
                p.get("total_variants"),
                p.get("main_image"),
                p.get("url"),
                p.get("created_at")[:10] if p.get("created_at") else ""
            ])

        # Apply currency & row formatting
        for row in ws_catalog.iter_rows(min_row=2, max_row=len(data) + 1):
            for col_idx, cell in enumerate(row, start=1):
                cell.font = regular_font
                cell.border = thin_border
                # Price columns (7 and 8)
                if col_idx in [7, 8] and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
                # Discount column (9)
                elif col_idx == 9 and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0"%"'
                # Center align status/gender/concentration
                elif col_idx in [4, 5, 6, 13, 14]:
                    cell.alignment = Alignment(horizontal="center")

        # ----------------------------------------------------
        # SHEET 2: VARIANTS BREAKDOWN
        # ----------------------------------------------------
        ws_variants = wb.create_sheet(title="Product Variants")
        variant_headers = ["Product ID", "Product Title", "Variant ID", "Variant Title", "SKU", "Price ($)", "Compare Price ($)", "In Stock"]
        ws_variants.append(variant_headers)

        for cell in ws_variants[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        variant_row_count = 1
        for p in data:
            for v in p.get("variants", []):
                ws_variants.append([
                    p.get("id"),
                    p.get("title"),
                    v.get("variant_id"),
                    v.get("variant_title"),
                    v.get("sku"),
                    v.get("price"),
                    v.get("compare_at_price"),
                    "YES" if v.get("available") else "NO"
                ])
                variant_row_count += 1

        for row in ws_variants.iter_rows(min_row=2, max_row=variant_row_count):
            for col_idx, cell in enumerate(row, start=1):
                cell.font = regular_font
                cell.border = thin_border
                if col_idx in [6, 7] and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
                elif col_idx in [8]:
                    cell.alignment = Alignment(horizontal="center")

        # ----------------------------------------------------
        # SHEET 3: CATALOG SUMMARY
        # ----------------------------------------------------
        ws_summary = wb.create_sheet(title="Executive Summary")
        ws_summary.append(["Lattafa Perfumes Catalog Analysis Report"])
        ws_summary.cell(row=1, column=1).font = title_font
        ws_summary.append([])

        total_products = len(data)
        prices = [p["min_price"] for p in data if p.get("min_price")]
        avg_price = sum(prices) / len(prices) if prices else 0.0
        in_stock_count = sum(1 for p in data if p.get("is_available"))

        summary_metrics = [
            ("Total Perfumes Scraped", total_products),
            ("In-Stock Perfumes", f"{in_stock_count} ({round(in_stock_count/total_products*100, 1) if total_products else 0}%)"),
            ("Average Price", f"${avg_price:.2f}"),
            ("Lowest Price", f"${min(prices):.2f}" if prices else "$0.00"),
            ("Highest Price", f"${max(prices):.2f}" if prices else "$0.00"),
        ]

        ws_summary.append(["Metric", "Value"])
        ws_summary.cell(row=3, column=1).font = header_font
        ws_summary.cell(row=3, column=1).fill = header_fill
        ws_summary.cell(row=3, column=2).font = header_font
        ws_summary.cell(row=3, column=2).fill = header_fill

        for metric, val in summary_metrics:
            ws_summary.append([metric, val])
            row_idx = ws_summary.max_row
            ws_summary.cell(row=row_idx, column=1).font = bold_font
            ws_summary.cell(row=row_idx, column=2).font = regular_font

        # Auto-adjust column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        wb.save(file_path)
        logging.info(f"Exported Excel Workbook to {file_path}")
        return file_path
